import os
from pprint import pprint
from typing import List, cast, Dict

import jirapt
from dotenv import load_dotenv
from jira import JIRA
from jira.client import ResultList
from jira.resources import Issue
from openpyxl import load_workbook
from funcoes import handling_fields

# - Carrega as variáveis de ambiente no documento .env
load_dotenv()


class JiraHandling:
    """
    Essa classe é responsável por criar a instancia do JIRA e fazer manipulaçãos nos métodos da LIB JIRA,

    create_jira_instance: Criar uma instancia do jira para ser utilizada por todos os outros métodos da clase;
    get_custom_fields: Esse metodo retorna uma lista com todos os campos do jira.
    set_custom_fields: Leio um documento em excel onde os usuários podem escolher os campos pelos nomes dos campos.
    get_issue_list: Retorna um Resultlist para montar um dicionário com os dados.

    """
    def __init__(self, usuario, api_token, servidor,jql):
        self._usuario = usuario
        self._api_token = api_token
        self._servidor = servidor
        self._jql = jql
        self._fields_to_dcit = {}
        self._instancia_jira = self.create_jira_instance()

    def create_jira_instance(self):
        """

        :return: Instância Jira.
        """
        jira = JIRA(basic_auth=(self._usuario, self._api_token),
                    server=self._servidor)

        myself = jira.myself()

        return jira

    def get_custom_fields(self) -> List:
        """

        :return:
        """
        fields: List = []
        try:
            fields = self._instancia_jira.fields()
        except Exception as error:
            print(f"Erro ao processar a request: {type(error).__name__} : - {error}")

        field_list: List = []
        for dicionario in fields:
            if dicionario.get('custom'):
                field_list.append((dicionario.get("name"), dicionario.get('key'), dicionario.get('schema').get("type")))
            else:
                field_list.append((dicionario.get("name"), dicionario.get('key')))

        return field_list

    def set_custom_fields_list(self) -> Dict:
        """
        Esse método realiza uma leitura em um excel onde estão os campos personalizados, escolidos pelo usuário.
        :return: Dict contendo os campos personalizados dentro do excel. (Resolução: resolution. )
        """
        try:
            fields_returned = self.get_custom_fields()
            wb = load_workbook(filename=f"{os.path.dirname(os.getcwd())}\\config\\conf.xlsx")
            ws = wb.active
            for row in ws.iter_rows(min_row=2, min_col=1):
                for field in fields_returned:
                    if field[0] == row[0].value:
                        self._fields_to_dcit[field[0]] = field[1]

            return self._fields_to_dcit
        except Exception as error:
            print(f"Erro ao processar a request: {type(error).__name__} : - {error}")

    def get_issue_list(self, jql) -> ResultList:
        """
        Esse método realiza uma query no jira, e retorna a lista de issues encontradas

        :param jql: String de pesquisa utilizado no Jira
        :return: Retorna uma lista (ResultList) com as issues.
        """
        try:
            dict_of_fields = self.set_custom_fields_list()
            issues_list = cast(ResultList[Issue], jirapt.search_issues(self._instancia_jira, jql, 2,
                                                                  fields=list(dict_of_fields.values())))
            return issues_list
        except Exception as error:
            print(f"Erro ao processar a request: {type(error).__name__} : - {error}")

    def create_dict_to_sharepoint(self) -> List:
        issues: ResultList = self.get_issue_list(self._jql)
        for issue in issues:
            pprint(issue.raw)
        dict_of_fields = self.set_custom_fields_list()
        list_of_issues: List = []
        for issue in issues:
            dict_issues: Dict = {}
            for value, name in zip(dict_of_fields.values(), dict_of_fields.keys()):
                dict_issues[name] = handling_fields(value, dict(issue.raw))
            list_of_issues.append(dict_issues)

        return list_of_issues


if __name__ == "__main__":
    jr = JiraHandling(os.environ['USER_JIRA'], os.environ['API_TOKEN'], os.environ['SERVIDOR'], os.environ['JQL'])
    issues = jr.create_dict_to_sharepoint()
    pprint(issues)